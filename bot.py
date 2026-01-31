import asyncio
import csv
import io
import os
from datetime import datetime
from typing import List, Set, Dict, Optional

from aiogram import Bot, Dispatcher, F
from aiogram.filters import CommandStart, Command
from aiogram.types import (
    Message,
    ReplyKeyboardMarkup,
    KeyboardButton,
    BufferedInputFile,
)
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from dotenv import load_dotenv
from openpyxl import load_workbook, Workbook


# ============ НАСТРОЙКИ ============

# Секретная команда для получения админ-доступа
ADMIN_SECRET_COMMAND = "get_bd_access_9876"

# Файл со списком админов (Telegram ID)
ADMINS_FILE = "admins.txt"

# Группа поддержки с топиками (Forum)
SUPPORT_GROUP_ID = -1003702935049

# Файл для хранения связей user_id -> topic_id
SUPPORT_TOPICS_FILE = "support_topics.csv"

# Типы баз данных и их настройки
# key: внутреннее имя, name: отображаемое имя, csv: файл, limit: лимит на пользователя
BASE_TYPES = {
    "telegram": {"name": "Telegram", "csv": "base_telegram.csv", "limit": 50},
    "whatsapp": {"name": "WhatsApp", "csv": "base_whatsapp.csv", "limit": 35},
    "max": {"name": "Max", "csv": "base_max.csv", "limit": 35},
    "viber": {"name": "Viber", "csv": "base_viber.csv", "limit": 35},
    "instagram": {"name": "Нельзяграм (там где Reels)", "csv": "base_instagram.csv", "limit": 300},
    "vk": {"name": "ВКонтакте", "csv": "base_vk.csv", "limit": 250},
    "ok": {"name": "Одноклассники", "csv": "base_ok.csv", "limit": 250},
    "email": {"name": "Почта", "csv": "base_email.csv", "limit": 100},
}

# Файл для хранения пользователей (счётчик)
USERS_FILE = "users.txt"

# Файл для хранения дополнительных лимитов (user_id, base_type, extra_limit)
USER_LIMITS_FILE = "user_limits.csv"

# Файл для хранения статусов пользователей (pending/approved/banned)
USER_STATUS_FILE = "user_status.csv"

# ID топика для заявок (создаётся автоматически или указать вручную)
REQUESTS_TOPIC_ID = None  # Будет создан автоматически

# Карта названий листов Excel -> внутренние ключи (для загрузки через админку)
EXCEL_SHEET_MAP = {
    # Короткие названия
    "Тг": "telegram",
    "ТГ": "telegram",
    "Вотсап": "whatsapp",
    "Макс": "max",
    "Вайбер": "viber",
    "Инст": "instagram",
    "ВК": "vk",
    "Ок": "ok",
    "Почта": "email",
    # Полные названия
    "Telegram": "telegram",
    "telegram": "telegram",
    "WhatsApp": "whatsapp",
    "Whatsapp": "whatsapp",
    "whatsapp": "whatsapp",
    "Max": "max",
    "max": "max",
    "Viber": "viber",
    "viber": "viber",
    "Нельзяграм": "instagram",
    "Нельзяграм (там где Reels)": "instagram",
    "Instagram": "instagram",
    "instagram": "instagram",
    "ВКонтакте": "vk",
    "Вконтакте": "vk",
    "вконтакте": "vk",
    "VK": "vk",
    "Одноклассники": "ok",
    "одноклассники": "ok",
    "OK": "ok",
    "Ok": "ok",
    "Email": "email",
    "email": "email",
    "Почты": "email",
}

# ============ НАЧАЛЬНАЯ ЗАГРУЗКА (ОТКЛЮЧЕНА) ============
# Раскомментируй для автозагрузки из файла при первом запуске:
# INITIAL_EXCEL_PATH = "Новая таблица.xlsx"
# INITIAL_LOAD_ENABLED = True


# ============ СОСТОЯНИЯ FSM ============

class AdminStates(StatesGroup):
    waiting_upload_choice = State()  # Ожидание выбора типа базы для загрузки
    waiting_file = State()  # Ожидание файла от админа
    waiting_delete_confirm = State()  # Ожидание подтверждения удаления базы


# ============ ГЛОБАЛЬНЫЕ ПЕРЕМЕННЫЕ ============

csv_lock = asyncio.Lock()
processing_users: Dict[str, Set[int]] = {key: set() for key in BASE_TYPES}


# ============ РАБОТА С АДМИНАМИ ============

def load_admins() -> Set[int]:
    """Загружает список админов из файла."""
    if not os.path.exists(ADMINS_FILE):
        return set()
    with open(ADMINS_FILE, "r", encoding="utf-8") as f:
        admins = set()
        for line in f:
            line = line.strip()
            if line:
                try:
                    admins.add(int(line))
                except ValueError:
                    pass
        return admins


def save_admin(user_id: int) -> None:
    """Добавляет админа в файл."""
    admins = load_admins()
    if user_id not in admins:
        with open(ADMINS_FILE, "a", encoding="utf-8") as f:
            f.write(f"{user_id}\n")


def is_admin(user_id: int) -> bool:
    """Проверяет, является ли пользователь админом."""
    return user_id in load_admins()


# ============ СЧЁТЧИК ПОЛЬЗОВАТЕЛЕЙ ============

def load_users() -> Set[int]:
    """Загружает список пользователей из файла."""
    if not os.path.exists(USERS_FILE):
        return set()
    with open(USERS_FILE, "r", encoding="utf-8") as f:
        users = set()
        for line in f:
            line = line.strip()
            if line:
                try:
                    users.add(int(line))
                except ValueError:
                    pass
        return users


def save_user(user_id: int) -> None:
    """Добавляет пользователя в файл (если ещё нет)."""
    users = load_users()
    if user_id not in users:
        with open(USERS_FILE, "a", encoding="utf-8") as f:
            f.write(f"{user_id}\n")


def get_users_count() -> int:
    """Возвращает количество пользователей."""
    return len(load_users())


# ============ ДОПОЛНИТЕЛЬНЫЕ ЛИМИТЫ ============

def load_user_limits() -> Dict[tuple, int]:
    """Загружает дополнительные лимиты: {(user_id, base_type): extra_limit}."""
    limits = {}
    if not os.path.exists(USER_LIMITS_FILE):
        return limits
    with open(USER_LIMITS_FILE, "r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        next(reader, None)  # Пропускаем заголовок
        for row in reader:
            if len(row) >= 3:
                try:
                    user_id = int(row[0])
                    base_type = row[1]
                    extra = int(row[2])
                    limits[(user_id, base_type)] = extra
                except ValueError:
                    pass
    return limits


def get_user_extra_limit(user_id: int, base_type: str) -> int:
    """Возвращает дополнительный лимит для пользователя по типу базы."""
    limits = load_user_limits()
    return limits.get((user_id, base_type), 0)


def set_user_extra_limit(user_id: int, base_type: str, value: int) -> None:
    """Устанавливает дополнительный лимит для пользователя."""
    limits = load_user_limits()
    key = (user_id, base_type)
    limits[key] = value
    
    # Сохраняем
    with open(USER_LIMITS_FILE, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["user_id", "base_type", "extra_limit"])
        for (uid, btype), extra in limits.items():
            writer.writerow([uid, btype, extra])


# ============ СТАТУСЫ ПОЛЬЗОВАТЕЛЕЙ ============
# Статусы: pending (ожидает), approved (одобрен), banned (забанен)

def load_user_statuses() -> Dict[int, str]:
    """Загружает статусы пользователей: {user_id: status}."""
    statuses = {}
    if not os.path.exists(USER_STATUS_FILE):
        return statuses
    with open(USER_STATUS_FILE, "r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        next(reader, None)
        for row in reader:
            if len(row) >= 2:
                try:
                    user_id = int(row[0])
                    status = row[1]
                    statuses[user_id] = status
                except ValueError:
                    pass
    return statuses


def get_user_status(user_id: int) -> Optional[str]:
    """Возвращает статус пользователя (pending/approved/banned) или None если не зарегистрирован."""
    statuses = load_user_statuses()
    return statuses.get(user_id)


def set_user_status(user_id: int, status: str) -> None:
    """Устанавливает статус пользователя."""
    statuses = load_user_statuses()
    statuses[user_id] = status
    
    with open(USER_STATUS_FILE, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["user_id", "status"])
        for uid, st in statuses.items():
            writer.writerow([uid, st])


def is_user_approved(user_id: int) -> bool:
    """Проверяет, одобрен ли пользователь."""
    return get_user_status(user_id) == "approved"


def is_user_banned(user_id: int) -> bool:
    """Проверяет, забанен ли пользователь."""
    return get_user_status(user_id) == "banned"


def is_user_pending(user_id: int) -> bool:
    """Проверяет, ожидает ли пользователь одобрения."""
    return get_user_status(user_id) == "pending"


# ============ РАБОТА С ТОПИКАМИ ПОДДЕРЖКИ ============

def load_support_topics() -> Dict[int, int]:
    """Загружает связи user_id -> topic_id из файла."""
    topics = {}
    if not os.path.exists(SUPPORT_TOPICS_FILE):
        return topics
    with open(SUPPORT_TOPICS_FILE, "r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        next(reader, None)  # Пропускаем заголовок
        for row in reader:
            if len(row) >= 2:
                try:
                    user_id = int(row[0])
                    topic_id = int(row[1])
                    topics[user_id] = topic_id
                except ValueError:
                    pass
    return topics


def save_support_topic(user_id: int, topic_id: int) -> None:
    """Сохраняет связь user_id -> topic_id."""
    topics = load_support_topics()
    topics[user_id] = topic_id
    
    with open(SUPPORT_TOPICS_FILE, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["user_id", "topic_id"])
        for uid, tid in topics.items():
            writer.writerow([uid, tid])


def get_user_by_topic(topic_id: int) -> Optional[int]:
    """Находит user_id по topic_id."""
    topics = load_support_topics()
    for uid, tid in topics.items():
        if tid == topic_id:
            return uid
    return None


# ============ ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ============

def clean_value(val) -> Optional[str]:
    """Убирает .0 у чисел, знак = в начале, возвращает строку."""
    if val is None:
        return None
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    s = str(val).strip()
    # Убираем знак = в начале (Excel иногда добавляет для формул)
    if s.startswith("="):
        s = s[1:]
    return s if s else None


def ensure_csv_exists() -> None:
    """Проверяет наличие CSV-файлов. Создаёт пустые, если нет."""
    for key, info in BASE_TYPES.items():
        csv_path = info["csv"]
        if not os.path.exists(csv_path):
            with open(csv_path, "w", encoding="utf-8", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["Value", "ID", "Username", "Date"])
            print(f"Создан пустой файл: {csv_path}")


# ============ РАБОТА С CSV ============

def _read_csv(path: str) -> List[List[str]]:
    """Читает CSV и возвращает список строк."""
    if not os.path.exists(path):
        return [["Value", "ID", "Username", "Date"]]
    with open(path, "r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        return list(reader)


def _write_csv(path: str, rows: List[List[str]]) -> None:
    """Записывает список строк в CSV."""
    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerows(rows)


def _count_user_records(rows: List[List[str]], user_id: int) -> int:
    """Считает, сколько записей выдано пользователю."""
    count = 0
    for row in rows[1:]:
        if len(row) > 1 and row[1]:
            try:
                if int(row[1]) == user_id:
                    count += 1
            except (ValueError, TypeError):
                continue
    return count


def _assign_records_csv(
    rows: List[List[str]],
    count: int,
    user_id: int,
    username: str,
) -> List[str]:
    """Берёт свободные записи, помечает как выданные."""
    taken: List[str] = []
    now = datetime.utcnow().strftime("%Y.%m.%d %H:%M:%S")

    for row in rows[1:]:
        if len(taken) >= count:
            break

        if len(row) < 4:
            row.extend([""] * (4 - len(row)))

        if row[1]:  # Уже выдано
            continue

        value = row[0].strip()
        # Убираем знак = в начале (Excel иногда добавляет)
        if value.startswith("="):
            value = value[1:]
        if not value:
            continue

        row[1] = str(user_id)
        row[2] = username or ""
        row[3] = now
        taken.append(value)

    return taken


def _get_existing_values(rows: List[List[str]]) -> Set[str]:
    """Возвращает множество всех значений в базе."""
    values = set()
    for row in rows[1:]:
        if row and row[0]:
            values.add(row[0].strip().lower())
    return values


def _add_new_values(csv_path: str, new_values: List[str]) -> int:
    """
    Добавляет новые значения в CSV с проверкой на дубликаты.
    Возвращает количество добавленных записей.
    """
    rows = _read_csv(csv_path)
    existing = _get_existing_values(rows)

    added = 0
    for val in new_values:
        val_clean = clean_value(val)
        if val_clean and val_clean.lower() not in existing:
            rows.append([val_clean, "", "", ""])
            existing.add(val_clean.lower())
            added += 1

    if added > 0:
        _write_csv(csv_path, rows)

    return added


# ============ ВЫДАЧА ДАННЫХ ============

async def allocate_for_user(base_key: str, user_id: int, username: str) -> tuple[List[str], str]:
    """
    Универсальная функция выдачи данных из любой базы.
    
    Возвращает кортеж: (список_контактов, причина_отказа)
    - причина: None — успех, "already_got" — уже получил, "not_enough" — недостаточно контактов
    """
    info = BASE_TYPES[base_key]
    csv_path = info["csv"]
    base_limit = info["limit"]
    
    # Учитываем дополнительный лимит от менеджера
    extra_limit = get_user_extra_limit(user_id, base_key)
    total_allowed = base_limit + extra_limit

    async with csv_lock:
        def _worker() -> tuple[List[str], str]:
            rows = _read_csv(csv_path)

            # Проверяем, сколько уже выдано этому пользователю
            current = _count_user_records(rows, user_id)
            if current >= total_allowed:
                return ([], "already_got")

            # Сколько ещё можно выдать
            can_give = total_allowed - current

            # Считаем свободные контакты (где нет ID)
            free_count = sum(1 for r in rows if len(r) < 2 or not r[1])
            if free_count < can_give:
                return ([], "not_enough")

            taken = _assign_records_csv(rows, can_give, user_id, username)

            if taken:
                _write_csv(csv_path, rows)

            return (taken, None)

        return await asyncio.to_thread(_worker)


# ============ СОЗДАНИЕ ФАЙЛОВ ============

def _create_txt_file(values: List[str], prefix: str) -> tuple[io.BytesIO, str]:
    """Создаёт txt-файл в памяти."""
    content = "\n".join(values)
    buffer = io.BytesIO(content.encode("utf-8"))
    filename = f"{prefix}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.txt"
    return buffer, filename


def _create_full_excel() -> tuple[io.BytesIO, str]:
    """Собирает все CSV-базы в один Excel-файл."""
    wb = Workbook()
    first = True

    for key, info in BASE_TYPES.items():
        csv_path = info["csv"]
        sheet_name = info["name"]

        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)

        if os.path.exists(csv_path):
            rows = _read_csv(csv_path)
            for row in rows:
                ws.append(row)
        else:
            ws.append(["Value", "ID", "Username", "Date"])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"full_base_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return buffer, filename


# ============ КЛАВИАТУРЫ ============

def get_main_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📦 Получить списки контактов")],
            [KeyboardButton(text="💬 Написать в поддержку")],
        ],
        resize_keyboard=True,
    )


def get_registration_keyboard() -> ReplyKeyboardMarkup:
    """Клавиатура для регистрации."""
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="✅ Отправить приглашение")],
        ],
        resize_keyboard=True,
    )


def get_user_choice_keyboard() -> ReplyKeyboardMarkup:
    """Клавиатура выбора типа контактов для пользователя."""
    return ReplyKeyboardMarkup(
        keyboard=[
            [
                KeyboardButton(text="📱 Telegram"),
                KeyboardButton(text="💬 WhatsApp"),
            ],
            [
                KeyboardButton(text="📨 Max"),
                KeyboardButton(text="📞 Viber"),
            ],
            [
                KeyboardButton(text="📷 Нельзяграм"),
                KeyboardButton(text="👥 ВКонтакте"),
            ],
            [
                KeyboardButton(text="🟠 Одноклассники"),
                KeyboardButton(text="📧 Почта"),
            ],
            [KeyboardButton(text="⬅️ Назад")],
        ],
        resize_keyboard=True,
    )


def get_admin_keyboard() -> ReplyKeyboardMarkup:
    """Клавиатура админа."""
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📥 Загрузить Базу данных")],
            [KeyboardButton(text="📤 Выкачать Базу данных")],
            [KeyboardButton(text="🗑 Удалить всю базу данных")],
            [KeyboardButton(text="⬅️ Выход из админки")],
        ],
        resize_keyboard=True,
    )


def get_delete_confirm_keyboard() -> ReplyKeyboardMarkup:
    """Клавиатура подтверждения удаления."""
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="✅ Да, удалить всё")],
            [KeyboardButton(text="❌ Отмена")],
        ],
        resize_keyboard=True,
    )


def get_admin_upload_choice_keyboard() -> ReplyKeyboardMarkup:
    """Клавиатура выбора типа базы для загрузки."""
    return ReplyKeyboardMarkup(
        keyboard=[
            [
                KeyboardButton(text="📱 Загрузить Telegram"),
                KeyboardButton(text="💬 Загрузить WhatsApp"),
            ],
            [
                KeyboardButton(text="📨 Загрузить Max"),
                KeyboardButton(text="📞 Загрузить Viber"),
            ],
            [
                KeyboardButton(text="📷 Загрузить Нельзяграм"),
                KeyboardButton(text="👥 Загрузить ВКонтакте"),
            ],
            [
                KeyboardButton(text="🟠 Загрузить Одноклассники"),
                KeyboardButton(text="📧 Загрузить Почта"),
            ],
            [KeyboardButton(text="📚 Загрузить ВСЕ листы из файла")],
            [KeyboardButton(text="⬅️ Отмена")],
        ],
        resize_keyboard=True,
    )


# ============ МАППИНГ КНОПОК ============

# Кнопки пользователя -> ключ базы
USER_BUTTON_MAP = {
    "📱 Telegram": "telegram",
    "💬 WhatsApp": "whatsapp",
    "📨 Max": "max",
    "📞 Viber": "viber",
    "📷 Нельзяграм": "instagram",
    "👥 ВКонтакте": "vk",
    "🟠 Одноклассники": "ok",
    "📧 Почта": "email",
}

# Кнопки админа для загрузки -> ключ базы
ADMIN_UPLOAD_MAP = {
    "📱 Загрузить Telegram": "telegram",
    "💬 Загрузить WhatsApp": "whatsapp",
    "📨 Загрузить Max": "max",
    "📞 Загрузить Viber": "viber",
    "📷 Загрузить Нельзяграм": "instagram",
    "👥 Загрузить ВКонтакте": "vk",
    "🟠 Загрузить Одноклассники": "ok",
    "📧 Загрузить Почта": "email",
    "📚 Загрузить ВСЕ листы из файла": "all",
}


# ============ ХЕНДЛЕРЫ ============

async def on_start(message: Message, state: FSMContext, bot: Bot) -> None:
    await state.clear()
    
    user = message.from_user
    if not user:
        return
    
    user_id = user.id
    status = get_user_status(user_id)
    
    # Сохраняем пользователя в счётчик
    save_user(user_id)
    
    # Проверяем статус пользователя
    if status == "banned":
        await message.answer(
            "🚫 Ваш аккаунт заблокирован.\n\n"
            "Обратитесь к администратору для разблокировки."
        )
        return
    
    if status == "approved":
        # Пользователь одобрен — показываем главное меню
        text = (
            "Привет!\n\n"
            "Этот бот выдаёт тебе списки контактов по которым нужно отправлять сообщения.\n\n"
            "Нажми кнопку ниже, затем выбери соц сеть или мессенджер где тебе удобнее работать."
        )
        await message.answer(text, reply_markup=get_main_keyboard())
        return
    
    if status == "pending":
        # Уже отправил заявку — ждёт одобрения
        await message.answer(
            "⏳ Ваша заявка уже отправлена!\n\n"
            "Ожидайте подтверждения от администратора."
        )
        return
    
    # Новый пользователь — показываем экран регистрации
    text = (
        "Если вы получили доступ к данному боту, значит вы уже прошли собеседование.\n\n"
        "Нажмите на кнопку ниже, админ примет приглашение и начнем ✅"
    )
    await message.answer(text, reply_markup=get_registration_keyboard())


async def on_send_request(message: Message, bot: Bot) -> None:
    """Пользователь нажал 'Отправить приглашение'."""
    user = message.from_user
    if not user:
        return
    
    # Проверяем что это личный чат
    if message.chat.type != "private":
        return
    
    user_id = user.id
    status = get_user_status(user_id)
    
    if status == "approved":
        await message.answer("Вы уже зарегистрированы!", reply_markup=get_main_keyboard())
        return
    
    if status == "pending":
        await message.answer("⏳ Ваша заявка уже отправлена! Ожидайте подтверждения.")
        return
    
    if status == "banned":
        await message.answer("🚫 Ваш аккаунт заблокирован.")
        return
    
    # Создаём заявку
    set_user_status(user_id, "pending")
    
    # Создаём топик для пользователя
    user_name = user.full_name or f"User {user_id}"
    if user.username:
        user_name += f" (@{user.username})"
    
    try:
        forum_topic = await bot.create_forum_topic(
            chat_id=SUPPORT_GROUP_ID,
            name=f"📝 {user_name[:120]}",
        )
        topic_id = forum_topic.message_thread_id
        save_support_topic(user_id, topic_id)
        
        # Отправляем заявку в топик
        await bot.send_message(
            chat_id=SUPPORT_GROUP_ID,
            message_thread_id=topic_id,
            text=(
                f"📝 НОВАЯ ЗАЯВКА!\n\n"
                f"👤 Пользователь: {user.full_name}\n"
                f"🆔 ID: {user_id}\n"
                f"📱 Username: @{user.username or 'нет'}\n\n"
                f"Для одобрения: /add\n"
                f"Для бана: /ban"
            ),
        )
        
        await message.answer(
            "✅ Заявка отправлена!\n\n"
            "Ожидайте подтверждения от администратора.\n"
            "Вам придёт уведомление когда заявка будет одобрена."
        )
    except Exception as e:
        set_user_status(user_id, None)  # Откатываем статус
        await message.answer(f"❌ Ошибка при отправке заявки: {e}")


async def on_add_user(message: Message, bot: Bot) -> None:
    """Команда /add — одобрить пользователя."""
    if message.chat.id != SUPPORT_GROUP_ID:
        return
    
    topic_id = message.message_thread_id
    if not topic_id:
        await message.answer("❌ Эта команда работает только в топике пользователя.")
        return
    
    user_id = get_user_by_topic(topic_id)
    if not user_id:
        await message.answer("❌ Пользователь не найден для этого топика.")
        return
    
    status = get_user_status(user_id)
    if status == "approved":
        await message.answer("ℹ️ Пользователь уже одобрен.")
        return
    
    set_user_status(user_id, "approved")
    await message.answer(f"✅ Пользователь {user_id} одобрен!")
    
    # Уведомляем пользователя
    try:
        await bot.send_message(
            chat_id=user_id,
            text=(
                "🎉 Ваша заявка одобрена!\n\n"
                "Теперь вы можете пользоваться ботом.\n"
                "Нажмите /start чтобы начать."
            ),
        )
    except Exception:
        pass


async def on_ban_user(message: Message, bot: Bot) -> None:
    """Команда /ban — забанить пользователя."""
    if message.chat.id != SUPPORT_GROUP_ID:
        return
    
    topic_id = message.message_thread_id
    if not topic_id:
        await message.answer("❌ Эта команда работает только в топике пользователя.")
        return
    
    user_id = get_user_by_topic(topic_id)
    if not user_id:
        await message.answer("❌ Пользователь не найден для этого топика.")
        return
    
    set_user_status(user_id, "banned")
    await message.answer(f"🚫 Пользователь {user_id} заблокирован!")
    
    # Уведомляем пользователя
    try:
        await bot.send_message(
            chat_id=user_id,
            text="🚫 Ваш аккаунт заблокирован.\n\nОбратитесь к администратору для разблокировки.",
        )
    except Exception:
        pass


async def on_unban_user(message: Message, bot: Bot) -> None:
    """Команда /unban — разбанить пользователя."""
    if message.chat.id != SUPPORT_GROUP_ID:
        return
    
    topic_id = message.message_thread_id
    if not topic_id:
        await message.answer("❌ Эта команда работает только в топике пользователя.")
        return
    
    user_id = get_user_by_topic(topic_id)
    if not user_id:
        await message.answer("❌ Пользователь не найден для этого топика.")
        return
    
    status = get_user_status(user_id)
    if status != "banned":
        await message.answer("ℹ️ Пользователь не заблокирован.")
        return
    
    set_user_status(user_id, "approved")
    await message.answer(f"✅ Пользователь {user_id} разблокирован!")
    
    # Уведомляем пользователя
    try:
        await bot.send_message(
            chat_id=user_id,
            text=(
                "✅ Ваш аккаунт разблокирован!\n\n"
                "Нажмите /start чтобы продолжить."
            ),
        )
    except Exception:
        pass


async def on_admin_command(message: Message, state: FSMContext) -> None:
    """Обработка секретной команды для получения админ-доступа."""
    # Только в личном чате с ботом
    if message.chat.type != "private":
        return
    
    user = message.from_user
    if not user:
        return

    save_admin(user.id)
    await state.clear()
    await message.answer(
        "✅ Админ-доступ активирован!\n\n"
        "Теперь тебе доступны функции управления базой данных.",
        reply_markup=get_admin_keyboard(),
    )


async def on_chatid(message: Message) -> None:
    """Показывает ID чата (для настройки группы поддержки)."""
    chat = message.chat
    topic_id = message.message_thread_id
    
    text = f"📍 **Информация о чате:**\n\n"
    text += f"Chat ID: `{chat.id}`\n"
    text += f"Тип: {chat.type}\n"
    if chat.title:
        text += f"Название: {chat.title}\n"
    if topic_id:
        text += f"Topic ID: `{topic_id}`\n"
    
    await message.answer(text, parse_mode="Markdown")


async def on_get_online(message: Message) -> None:
    """Показывает количество пользователей бота (только для группы админов)."""
    # Только в группе поддержки
    if message.chat.id != SUPPORT_GROUP_ID:
        return
    
    count = get_users_count()
    await message.answer(
        f"📊 Статистика бота:\n\n"
        f"👥 Всего пользователей: {count}"
    )


async def on_download_db(message: Message) -> None:
    """Выгрузка всей базы данных (только для группы админов)."""
    # Только в группе поддержки
    if message.chat.id != SUPPORT_GROUP_ID:
        return
    
    await message.answer("⏳ Собираю базу данных...")
    
    try:
        file_buffer, filename = await asyncio.to_thread(_create_full_excel)
        document = BufferedInputFile(file_buffer.read(), filename=filename)
        await message.answer_document(
            document=document,
            caption="📤 Полная база данных"
        )
    except Exception as e:
        await message.answer(f"❌ Ошибка при выгрузке: {e}")


async def on_stats(message: Message) -> None:
    """Статистика свободных контактов и выданных за периоды (только для группы админов)."""
    # Только в группе поддержки
    if message.chat.id != SUPPORT_GROUP_ID:
        return
    
    def _count_stats() -> tuple:
        from datetime import timedelta
        
        now = datetime.utcnow()
        day_ago = now - timedelta(days=1)
        week_ago = now - timedelta(days=7)
        month_ago = now - timedelta(days=30)
        
        free_stats = []  # (name, free, total)
        issued_stats = []  # (name, day, week, month)
        
        for key, info in BASE_TYPES.items():
            csv_path = info["csv"]
            rows = _read_csv(csv_path)
            total = len(rows) - 1  # Минус заголовок
            free = sum(1 for r in rows[1:] if len(r) < 2 or not r[1])
            free_stats.append((info["name"], free, total))
            
            # Считаем выданные за периоды
            day_count = 0
            week_count = 0
            month_count = 0
            
            for row in rows[1:]:
                if len(row) >= 4 and row[3]:  # Есть дата выдачи
                    try:
                        # Формат: "YYYY.MM.DD HH:MM:SS"
                        issued_date = datetime.strptime(row[3], "%Y.%m.%d %H:%M:%S")
                        if issued_date >= day_ago:
                            day_count += 1
                        if issued_date >= week_ago:
                            week_count += 1
                        if issued_date >= month_ago:
                            month_count += 1
                    except ValueError:
                        pass
            
            issued_stats.append((info["name"], day_count, week_count, month_count))
        
        return free_stats, issued_stats
    
    free_stats, issued_stats = await asyncio.to_thread(_count_stats)
    
    # Свободные контакты
    lines = ["📊 **Свободные контакты:**\n"]
    total_free = 0
    total_all = 0
    
    for name, free, total in free_stats:
        if free == 0:
            status = "🔴"
        elif free < 100:
            status = "🟡"
        else:
            status = "🟢"
        lines.append(f"{status} **{name}**: {free} / {total}")
        total_free += free
        total_all += total
    
    lines.append(f"\n📦 **Итого**: {total_free} свободных / {total_all} всего")
    
    # Выданные за периоды
    lines.append("\n\n📈 **Выдано контактов:**\n")
    lines.append("```")
    lines.append(f"{'Тип':<25} {'Сутки':>7} {'Неделя':>7} {'Месяц':>7}")
    lines.append("-" * 48)
    
    total_day = 0
    total_week = 0
    total_month = 0
    
    for name, day, week, month in issued_stats:
        # Обрезаем длинные названия
        short_name = name[:24] if len(name) > 24 else name
        lines.append(f"{short_name:<25} {day:>7} {week:>7} {month:>7}")
        total_day += day
        total_week += week
        total_month += month
    
    lines.append("-" * 48)
    lines.append(f"{'ИТОГО':<25} {total_day:>7} {total_week:>7} {total_month:>7}")
    lines.append("```")
    
    await message.answer("\n".join(lines), parse_mode="Markdown")


async def on_get_base(message: Message, state: FSMContext) -> None:
    user = message.from_user
    if not user or not is_user_approved(user.id):
        await message.answer("❌ У вас нет доступа к этой функции.")
        return
    
    await state.clear()
    await message.answer(
        "Выбери, какую базу хочешь получить:",
        reply_markup=get_user_choice_keyboard(),
    )


async def on_back(message: Message, state: FSMContext, bot: Bot) -> None:
    await state.clear()
    await on_start(message, state, bot)


async def on_user_base_choice(message: Message, state: FSMContext, bot: Bot) -> None:
    """Обработка выбора типа базы пользователем."""
    user = message.from_user
    if not user:
        await message.answer("Не удалось определить пользователя.")
        return
    
    if not is_user_approved(user.id):
        await message.answer("❌ У вас нет доступа к этой функции.")
        return

    text = message.text
    if text not in USER_BUTTON_MAP:
        return

    base_key = USER_BUTTON_MAP[text]
    info = BASE_TYPES[base_key]
    user_id = user.id
    username = user.username or ""

    # Защита от повторных нажатий
    if user_id in processing_users[base_key]:
        await message.answer("Подожди, твой запрос уже обрабатывается...")
        return

    processing_users[base_key].add(user_id)
    try:
        values, reason = await allocate_for_user(base_key, user_id, username)
    except Exception:
        await message.answer("Произошла ошибка. Попробуй позже.")
        return
    finally:
        processing_users[base_key].discard(user_id)

    if reason == "already_got":
        await message.answer(
            f"Ты уже получил контакты из «{info['name']}».\n"
            f"Лимит: {info['limit']} контактов. Обратитесь к менеджеру."
        )
        return

    if reason == "not_enough":
        await message.answer(
            f"❌ К сожалению, контакты «{info['name']}» на данный момент отсутствуют.\n"
            "Обратитесь к менеджеру."
        )
        # Уведомление в General (группу поддержки)
        try:
            await bot.send_message(
                chat_id=SUPPORT_GROUP_ID,
                text=(
                    f"⚠️ ВНИМАНИЕ: Контакты закончились!\n\n"
                    f"📦 Тип: {info['name']}\n"
                    f"👤 Пользователь: {user.full_name} (@{user.username or 'нет'})\n"
                    f"🆔 ID: {user_id}\n\n"
                    f"Необходимо загрузить новые контакты!"
                ),
            )
        except Exception:
            pass  # Не прерываем, если не удалось отправить
        return

    if not values:
        await message.answer("Произошла ошибка при выдаче контактов.")
        return

    # Отправляем контакты сообщением (не файлом)
    contacts_text = "\n".join(values)
    
    # Telegram ограничивает длину сообщения 4096 символами
    if len(contacts_text) <= 4000:
        await message.answer(
            f"✅ Выдано из «{info['name']}»: {len(values)} контактов\n\n"
            f"{contacts_text}"
        )
    else:
        # Если слишком длинный, разбиваем на части
        await message.answer(f"✅ Выдано из «{info['name']}»: {len(values)} контактов")
        
        # Отправляем по частям
        chunk = ""
        for val in values:
            if len(chunk) + len(val) + 1 > 4000:
                await message.answer(chunk)
                chunk = val
            else:
                chunk = chunk + "\n" + val if chunk else val
        if chunk:
            await message.answer(chunk)
    
    # Проверяем, осталось ли меньше 5% свободных контактов
    try:
        csv_path = info["csv"]
        rows = _read_csv(csv_path)
        total = len(rows) - 1  # Минус заголовок
        free = sum(1 for r in rows[1:] if len(r) < 2 or not r[1])
        
        if total > 0:
            percent = (free / total) * 100
            if percent < 5:
                await bot.send_message(
                    chat_id=SUPPORT_GROUP_ID,
                    text=(
                        f"⚠️ ВНИМАНИЕ: Контакты заканчиваются!\n\n"
                        f"📦 Тип: {info['name']}\n"
                        f"📊 Осталось: {free} из {total} ({percent:.1f}%)\n\n"
                        f"Необходимо загрузить новые контакты!"
                    ),
                )
    except Exception:
        pass


# ============ АДМИН-ХЕНДЛЕРЫ ============

async def on_admin_exit(message: Message, state: FSMContext) -> None:
    """Выход из админки."""
    await state.clear()
    await message.answer("Вышел из админки.", reply_markup=get_main_keyboard())


async def on_admin_download(message: Message, state: FSMContext) -> None:
    """Выкачивание всей базы в Excel."""
    user = message.from_user
    if not user or not is_admin(user.id):
        return

    try:
        buffer, filename = _create_full_excel()
    except Exception:
        await message.answer("Ошибка при создании файла.")
        return

    document = BufferedInputFile(buffer.read(), filename=filename)
    await message.answer_document(
        document=document,
        caption="📤 Полная база данных (все типы) в одном Excel-файле.",
    )


async def on_admin_upload_start(message: Message, state: FSMContext) -> None:
    """Начало загрузки базы — показываем выбор типа."""
    user = message.from_user
    if not user or not is_admin(user.id):
        return

    await state.set_state(AdminStates.waiting_upload_choice)
    await message.answer(
        "Выбери, какую базу хочешь загрузить:",
        reply_markup=get_admin_upload_choice_keyboard(),
    )


async def on_admin_upload_cancel(message: Message, state: FSMContext) -> None:
    """Отмена загрузки."""
    await state.clear()
    await message.answer("Загрузка отменена.", reply_markup=get_admin_keyboard())


async def on_admin_upload_choice(message: Message, state: FSMContext) -> None:
    """Обработка выбора типа базы для загрузки."""
    user = message.from_user
    if not user or not is_admin(user.id):
        return

    text = message.text
    if text not in ADMIN_UPLOAD_MAP:
        return

    base_key = ADMIN_UPLOAD_MAP[text]
    await state.update_data(upload_type=base_key)
    await state.set_state(AdminStates.waiting_file)

    if base_key == "all":
        await message.answer(
            "📚 Режим загрузки ВСЕХ листов.\n\n"
            "Отправь Excel-файл (.xlsx) с листами:\n"
            "Тг, Вотсап, Макс, Вайбер, Инст, ВК, Ок, Почта\n\n"
            "Данные будут добавлены в соответствующие базы без дубликатов.",
            reply_markup=ReplyKeyboardMarkup(
                keyboard=[[KeyboardButton(text="⬅️ Отмена")]],
                resize_keyboard=True,
            ),
        )
    else:
        info = BASE_TYPES[base_key]
        await message.answer(
            f"📥 Загрузка в базу «{info['name']}»\n\n"
            "Отправь Excel-файл (.xlsx).\n"
            "Данные будут взяты из первого столбца первого листа.\n"
            "Дубликаты автоматически пропускаются.",
            reply_markup=ReplyKeyboardMarkup(
                keyboard=[[KeyboardButton(text="⬅️ Отмена")]],
                resize_keyboard=True,
            ),
        )


async def on_admin_file_received(message: Message, state: FSMContext, bot: Bot) -> None:
    """Обработка полученного файла от админа."""
    user = message.from_user
    if not user or not is_admin(user.id):
        return

    if not message.document:
        await message.answer("Пожалуйста, отправь файл Excel (.xlsx).")
        return

    filename = message.document.file_name or ""
    if not filename.lower().endswith(".xlsx"):
        await message.answer("Нужен файл в формате .xlsx (Excel).")
        return

    data = await state.get_data()
    upload_type = data.get("upload_type")
    if not upload_type:
        await message.answer("Ошибка состояния. Начни заново.")
        await state.clear()
        return

    # Скачиваем файл
    await message.answer("⏳ Обрабатываю файл...")

    try:
        file_io = await bot.download(message.document)
        if not file_io:
            await message.answer("Не удалось скачать файл.")
            return

        wb = load_workbook(file_io, read_only=True)

        results = []

        if upload_type == "all":
            # Обрабатываем все листы
            for sheet_name in wb.sheetnames:
                base_key = EXCEL_SHEET_MAP.get(sheet_name)
                if not base_key:
                    results.append(f"⚠️ Лист «{sheet_name}» — неизвестный тип, пропущен")
                    continue

                ws = wb[sheet_name]
                new_values = []
                for row in ws.iter_rows(min_row=2, values_only=True):  # min_row=2 — пропускаем заголовок
                    val = clean_value(row[0] if row else None)
                    if val and val.lower() not in ("value", "значение", "контакт", "данные"):
                        new_values.append(val)

                if new_values:
                    csv_path = BASE_TYPES[base_key]["csv"]
                    added = _add_new_values(csv_path, new_values)
                    info = BASE_TYPES[base_key]
                    results.append(
                        f"✅ «{info['name']}» — добавлено {added} из {len(new_values)}"
                    )
                else:
                    results.append(f"⚠️ Лист «{sheet_name}» — пустой")
        else:
            # Обрабатываем первый лист для конкретного типа
            ws = wb.active
            new_values = []
            for row in ws.iter_rows(min_row=2, values_only=True):  # min_row=2 — пропускаем заголовок
                val = clean_value(row[0] if row else None)
                if val and val.lower() not in ("value", "значение", "контакт", "данные"):
                    new_values.append(val)

            if new_values:
                csv_path = BASE_TYPES[upload_type]["csv"]
                added = _add_new_values(csv_path, new_values)
                info = BASE_TYPES[upload_type]
                results.append(
                    f"✅ «{info['name']}» — добавлено {added} из {len(new_values)} "
                    f"(дубликатов пропущено: {len(new_values) - added})"
                )
            else:
                results.append("⚠️ Файл пустой или не содержит данных в первом столбце")

        wb.close()

        await state.clear()
        await message.answer(
            "📊 Результат загрузки:\n\n" + "\n".join(results),
            reply_markup=get_admin_keyboard(),
        )

    except Exception as e:
        await message.answer(f"❌ Ошибка при обработке файла: {e}")
        await state.clear()


# ============ АДМИН: УДАЛЕНИЕ БАЗЫ ============

def clear_all_databases() -> int:
    """Очищает все CSV-файлы (удаляет данные, оставляет заголовки). Возвращает кол-во очищенных."""
    count = 0
    for key, info in BASE_TYPES.items():
        csv_path = info["csv"]
        if os.path.exists(csv_path):
            with open(csv_path, "w", encoding="utf-8", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["Value", "ID", "Username", "Date"])
            count += 1
    return count


async def on_admin_delete_start(message: Message, state: FSMContext) -> None:
    """Начало удаления базы — показываем подтверждение."""
    user = message.from_user
    if not user or not is_admin(user.id):
        return

    await state.set_state(AdminStates.waiting_delete_confirm)
    await message.answer(
        "⚠️ ВНИМАНИЕ!\n\n"
        "Ты собираешься удалить ВСЮ базу данных.\n"
        "Это действие НЕОБРАТИМО!\n\n"
        "Все записи во всех типах баз будут удалены.\n\n"
        "Ты уверен?",
        reply_markup=get_delete_confirm_keyboard(),
    )


async def on_admin_delete_confirm(message: Message, state: FSMContext) -> None:
    """Подтверждение удаления."""
    user = message.from_user
    if not user or not is_admin(user.id):
        return

    async with csv_lock:
        count = await asyncio.to_thread(clear_all_databases)

    await state.clear()
    await message.answer(
        f"🗑 База данных полностью очищена.\n"
        f"Очищено файлов: {count}",
        reply_markup=get_admin_keyboard(),
    )


async def on_admin_delete_cancel(message: Message, state: FSMContext) -> None:
    """Отмена удаления."""
    await state.clear()
    await message.answer(
        "❌ Удаление отменено.",
        reply_markup=get_admin_keyboard(),
    )


# ============ ПОДДЕРЖКА: ХЕНДЛЕРЫ ============

async def on_support_info(message: Message) -> None:
    """Пользователь нажал 'Написать в поддержку' — показываем информацию."""
    await message.answer(
        "💬 Чтобы связаться с поддержкой, просто напиши любое сообщение в этот чат.\n\n"
        "Твоё сообщение будет отправлено менеджеру, и он ответит тебе здесь."
    )


async def on_user_message_to_support(message: Message, bot: Bot) -> None:
    """Любое сообщение от пользователя пересылается в поддержку."""
    # Только личные чаты
    if message.chat.type != "private":
        return
    
    user = message.from_user
    if not user:
        return
    
    # Проверяем статус пользователя
    if not is_user_approved(user.id):
        # Если пользователь не одобрен — не пересылаем
        return

    topics = load_support_topics()
    topic_id = topics.get(user.id)

    async def create_new_topic():
        """Создаёт новый топик для пользователя."""
        user_name = user.full_name or f"User {user.id}"
        if user.username:
            user_name += f" (@{user.username})"

        forum_topic = await bot.create_forum_topic(
            chat_id=SUPPORT_GROUP_ID,
            name=user_name[:128],
        )
        new_topic_id = forum_topic.message_thread_id
        save_support_topic(user.id, new_topic_id)

        # Приветственное сообщение в топик
        await bot.send_message(
            chat_id=SUPPORT_GROUP_ID,
            message_thread_id=new_topic_id,
            text=(
                f"🆕 Новый диалог!\n\n"
                f"👤 Пользователь: {user.full_name}\n"
                f"🆔 ID: {user.id}\n"
                f"📱 Username: @{user.username or 'нет'}"
            ),
        )
        return new_topic_id

    # Если топика нет — создаём
    if not topic_id:
        try:
            topic_id = await create_new_topic()
        except Exception as e:
            await message.answer(f"❌ Не удалось создать чат с поддержкой: {e}")
            return

    try:
        # Пересылаем сообщение в топик
        await message.forward(
            chat_id=SUPPORT_GROUP_ID,
            message_thread_id=topic_id,
        )
        await message.answer("✅ Сообщение отправлено в поддержку.")
    except Exception as e:
        # Если топик удалён — пересоздаём
        if "thread not found" in str(e).lower() or "message thread not found" in str(e).lower():
            try:
                topic_id = await create_new_topic()
                await message.forward(
                    chat_id=SUPPORT_GROUP_ID,
                    message_thread_id=topic_id,
                )
                await message.answer("✅ Сообщение отправлено в поддержку.")
            except Exception as e2:
                await message.answer(f"❌ Не удалось отправить сообщение: {e2}")
        else:
            await message.answer(f"❌ Не удалось отправить сообщение: {e}")


async def on_support_admin_reply(message: Message, bot: Bot) -> None:
    """Админ ответил в топике — отправляем пользователю от имени бота."""
    # Проверяем, что это сообщение из группы поддержки
    if message.chat.id != SUPPORT_GROUP_ID:
        return

    # Проверяем, что это ответ в топике (не в General)
    topic_id = message.message_thread_id
    if not topic_id:
        return

    # Игнорируем сообщения от бота
    if message.from_user and message.from_user.is_bot:
        return

    # Находим пользователя по topic_id
    user_id = get_user_by_topic(topic_id)
    if not user_id:
        return

    try:
        # Отправляем сообщение от имени бота (без пересылки)
        if message.text:
            await bot.send_message(
                chat_id=user_id,
                text=f"💬 Поддержка:\n\n{message.text}",
            )
        elif message.photo:
            await bot.send_photo(
                chat_id=user_id,
                photo=message.photo[-1].file_id,
                caption=f"💬 Поддержка:\n\n{message.caption or ''}",
            )
        elif message.document:
            await bot.send_document(
                chat_id=user_id,
                document=message.document.file_id,
                caption=f"💬 Поддержка:\n\n{message.caption or ''}",
            )
        elif message.voice:
            await bot.send_voice(
                chat_id=user_id,
                voice=message.voice.file_id,
                caption="💬 Голосовое от поддержки",
            )
        elif message.video:
            await bot.send_video(
                chat_id=user_id,
                video=message.video.file_id,
                caption=f"💬 Поддержка:\n\n{message.caption or ''}",
            )
        elif message.sticker:
            await bot.send_sticker(
                chat_id=user_id,
                sticker=message.sticker.file_id,
            )
    except Exception:
        # Пользователь мог заблокировать бота
        pass


# ============ МЕНЕДЖЕР: РАЗБЛОКИРОВКА ЛИМИТОВ ============

def get_user_used_types(user_id: int) -> List[str]:
    """Возвращает список типов баз, которые пользователь уже получал."""
    used = []
    for key, info in BASE_TYPES.items():
        csv_path = info["csv"]
        rows = _read_csv(csv_path)
        count = _count_user_records(rows, user_id)
        if count > 0:
            used.append(key)
    return used


def get_user_contacts(user_id: int) -> Dict[str, List[str]]:
    """Возвращает все контакты, выданные пользователю, по типам."""
    result = {}
    for key, info in BASE_TYPES.items():
        csv_path = info["csv"]
        rows = _read_csv(csv_path)
        contacts = []
        for row in rows[1:]:  # Пропускаем заголовок
            if len(row) >= 2 and row[1] == str(user_id):
                value = row[0].strip()
                if value.startswith("="):
                    value = value[1:]
                if value:
                    contacts.append(value)
        if contacts:
            result[key] = contacts
    return result


def _create_user_contacts_excel(user_id: int, contacts: Dict[str, List[str]]) -> tuple[io.BytesIO, str]:
    """Создаёт Excel-файл с контактами пользователя."""
    wb = Workbook()
    first = True
    
    for key, values in contacts.items():
        info = BASE_TYPES[key]
        if first:
            ws = wb.active
            ws.title = info["name"][:31]  # Максимум 31 символ для названия листа
            first = False
        else:
            ws = wb.create_sheet(title=info["name"][:31])
        
        ws.append(["Контакт"])
        for val in values:
            ws.append([val])
    
    if first:
        # Нет контактов — пустой файл
        ws = wb.active
        ws.title = "Пусто"
        ws.append(["Нет выданных контактов"])
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"contacts_user_{user_id}.xlsx"
    return buffer, filename


async def on_contacts_command(message: Message) -> None:
    """Команда /contacts в топике — показать выданные контакты пользователю."""
    # Только в группе поддержки
    if message.chat.id != SUPPORT_GROUP_ID:
        return
    
    topic_id = message.message_thread_id
    if not topic_id:
        await message.answer("❌ Эта команда работает только в топике пользователя.")
        return
    
    # Находим пользователя по топику
    user_id = get_user_by_topic(topic_id)
    if not user_id:
        await message.answer("❌ Пользователь не найден для этого топика.")
        return
    
    await message.answer("⏳ Собираю контакты пользователя...")
    
    # Получаем контакты
    contacts = await asyncio.to_thread(get_user_contacts, user_id)
    
    if not contacts:
        await message.answer("ℹ️ Этому пользователю ещё не выдавались контакты.")
        return
    
    # Формируем статистику
    stats = []
    total = 0
    for key, values in contacts.items():
        info = BASE_TYPES[key]
        stats.append(f"• {info['name']}: {len(values)}")
        total += len(values)
    
    # Создаём Excel
    file_buffer, filename = await asyncio.to_thread(
        _create_user_contacts_excel, user_id, contacts
    )
    document = BufferedInputFile(file_buffer.read(), filename=filename)
    
    await message.answer_document(
        document=document,
        caption=(
            f"📋 Контакты пользователя {user_id}:\n\n"
            + "\n".join(stats) +
            f"\n\n📊 Всего: {total} контактов"
        ),
    )


async def on_clear_command(message: Message, bot: Bot) -> None:
    """Команда /clear в топике — автоматическая разблокировка лимитов."""
    # Проверяем, что это в группе поддержки
    if message.chat.id != SUPPORT_GROUP_ID:
        return
    
    topic_id = message.message_thread_id
    if not topic_id:
        await message.answer("❌ Эта команда работает только в топике пользователя.")
        return
    
    # Находим пользователя по топику
    user_id = get_user_by_topic(topic_id)
    if not user_id:
        await message.answer("❌ Пользователь не найден для этого топика.")
        return
    
    # Находим типы, которые пользователь уже получал
    used_types = await asyncio.to_thread(get_user_used_types, user_id)
    
    if not used_types:
        await message.answer("ℹ️ Пользователь ещё не получал никаких контактов.")
        return
    
    # Разблокируем только те типы, где пользователь использовал весь лимит
    unlocked = []
    unlocked_keys = []
    skipped = []
    
    for key in used_types:
        info = BASE_TYPES[key]
        base_limit = info["limit"]
        extra_limit = get_user_extra_limit(user_id, key)
        total_allowed = base_limit + extra_limit
        
        # Считаем сколько уже получил
        csv_path = info["csv"]
        rows = _read_csv(csv_path)
        current = _count_user_records(rows, user_id)
        
        # Разблокируем только если использовал весь текущий лимит
        if current >= total_allowed:
            # Устанавливаем extra_limit = current, чтобы можно было получить ещё base_limit
            set_user_extra_limit(user_id, key, current)
            unlocked.append(f"• {info['name']} (+{base_limit})")
            unlocked_keys.append(key)
        else:
            remaining = total_allowed - current
            skipped.append(f"• {info['name']} (осталось {remaining})")
    
    if not unlocked:
        await message.answer("ℹ️ Пользователь ещё не использовал текущий лимит.")
        return
    
    await message.answer(f"✅ Разблокировано для пользователя:\n\n" + "\n".join(unlocked))
    
    # Уведомляем пользователя
    try:
        unlocked_names = [BASE_TYPES[k]["name"] for k in unlocked_keys]
        await bot.send_message(
            chat_id=user_id,
            text=(
                "🎉 Менеджер разблокировал тебе контакты!\n\n"
                f"Разблокировано: {', '.join(unlocked_names)}\n\n"
                "Теперь ты можешь получить ещё одну порцию."
            ),
        )
    except Exception:
        pass


# ============ ЗАПУСК ============

async def main() -> None:
    load_dotenv()
    token = os.getenv("BOT_TOKEN")
    if not token:
        raise RuntimeError("Не задан BOT_TOKEN в .env файле")

    # Создаём CSV если нужно
    ensure_csv_exists()

    bot = Bot(token=token)
    storage = MemoryStorage()
    dp = Dispatcher(storage=storage)

    # Базовые команды
    dp.message.register(on_start, CommandStart())
    dp.message.register(on_admin_command, Command(ADMIN_SECRET_COMMAND))
    dp.message.register(on_chatid, Command("chatid"))
    dp.message.register(on_get_online, Command("get_online"))
    dp.message.register(on_download_db, Command("download_db"))
    dp.message.register(on_stats, Command("stats"))
    
    # Регистрация пользователя
    dp.message.register(on_send_request, F.text == "✅ Отправить приглашение")
    
    # Команды модерации (в группе поддержки)
    dp.message.register(on_add_user, Command("add"), F.chat.id == SUPPORT_GROUP_ID)
    dp.message.register(on_ban_user, Command("ban"), F.chat.id == SUPPORT_GROUP_ID)
    dp.message.register(on_unban_user, Command("unban"), F.chat.id == SUPPORT_GROUP_ID)

    # Админ: состояние ожидания файла (должно быть выше остальных!)
    dp.message.register(
        on_admin_file_received,
        AdminStates.waiting_file,
        F.document,
    )
    dp.message.register(
        on_admin_upload_cancel,
        AdminStates.waiting_file,
        F.text == "⬅️ Отмена",
    )

    # Админ: состояние выбора типа загрузки
    dp.message.register(
        on_admin_upload_cancel,
        AdminStates.waiting_upload_choice,
        F.text == "⬅️ Отмена",
    )
    dp.message.register(
        on_admin_upload_choice,
        AdminStates.waiting_upload_choice,
    )

    # Админ: состояние подтверждения удаления
    dp.message.register(
        on_admin_delete_confirm,
        AdminStates.waiting_delete_confirm,
        F.text == "✅ Да, удалить всё",
    )
    dp.message.register(
        on_admin_delete_cancel,
        AdminStates.waiting_delete_confirm,
        F.text == "❌ Отмена",
    )

    # Менеджер: команда /clear в группе поддержки (ДО on_support_admin_reply!)
    dp.message.register(
        on_clear_command,
        Command("clear"),
        F.chat.id == SUPPORT_GROUP_ID,
    )
    
    # Менеджер: команда /contacts в группе поддержки (ДО on_support_admin_reply!)
    dp.message.register(
        on_contacts_command,
        Command("contacts"),
        F.chat.id == SUPPORT_GROUP_ID,
    )

    # Поддержка: ответы админов из группы (только supergroup, не личные чаты)
    dp.message.register(
        on_support_admin_reply,
        F.chat.type == "supergroup",
        F.chat.id == SUPPORT_GROUP_ID,
        ~Command("clear"),
        ~Command("contacts"),
        ~Command("add"),
        ~Command("ban"),
        ~Command("unban"),
    )
    
    # Админ: основные кнопки
    dp.message.register(on_admin_download, F.text == "📤 Выкачать Базу данных")
    dp.message.register(on_admin_upload_start, F.text == "📥 Загрузить Базу данных")
    dp.message.register(on_admin_delete_start, F.text == "🗑 Удалить всю базу данных")
    dp.message.register(on_admin_exit, F.text == "⬅️ Выход из админки")

    # Пользователь: навигация
    dp.message.register(on_get_base, F.text == "📦 Получить списки контактов")
    dp.message.register(on_support_info, F.text == "💬 Написать в поддержку")
    dp.message.register(on_back, F.text == "⬅️ Назад")

    # Пользователь: выбор типа базы
    for btn_text in USER_BUTTON_MAP:
        dp.message.register(on_user_base_choice, F.text == btn_text)

    # Все остальные сообщения в личном чате -> поддержка (ПОСЛЕДНИЙ хендлер!)
    dp.message.register(on_user_message_to_support, F.chat.type == "private")

    print("Бот запущен!")
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
